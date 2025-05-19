import os
import re
import pandas as pd
import fitz  # PyMuPDF for handling PDF files
from concurrent.futures import ThreadPoolExecutor
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from ttkbootstrap import Style, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def limpar_cnpj(cnpj):
    """
    Remove todos os caracteres não numéricos de um CNPJ.
    """
    return re.sub(r'\D', '', str(cnpj))


def extrair_comp(texto):
    """
    Extrai mês e ano de competência a partir do texto.
    """
    match = re.search(r'COMP\s*[:\.\-]?\s*(\d{2})/(\d{4})', texto)
    return (match.group(1), match.group(2)) if match else (None, None)


def coletar_pastas(pasta_base):
    """
    Retorna lista de tuplas (ano, mes_str, caminho_pasta) para todas pastas mês existentes.
    """
    pastas = []
    for ano in os.listdir(pasta_base):
        dir_ano = os.path.join(pasta_base, ano)
        if not os.path.isdir(dir_ano):
            continue
        for mes in range(1, 13):
            mes_str = f"{mes:02d}_{ano}"
            path = os.path.join(dir_ano, mes_str)
            if os.path.isdir(path):
                pastas.append((ano, mes_str, path))
    return pastas


def carregar_textos_pdfs(pdfs):
    """
    Lê todas as páginas de cada PDF uma única vez em cache.
    Retorna dict: {pdf_path: [texto_página0, texto_página1, ...]}
    """
    cache = {}
    for p in pdfs:
        try:
            doc = fitz.open(p)
            cache[p] = [doc.load_page(i).get_text() or "" for i in range(len(doc))]
            doc.close()
        except Exception:
            cache[p] = []
    return cache


def processar(caminho_excel, pasta_pdf_base, pasta_destino, status_text, progress_callback, pause_event, terminal):
    try:
        os.makedirs(pasta_destino, exist_ok=True)
        status_text.set("📚 Lendo lista de CNPJs...")
        terminal.insert(tk.END, "📚 Lendo lista de CNPJs...\n")
        terminal.see(tk.END)

        # Carrega Excel
        df = pd.read_excel(caminho_excel)
        cnpjs = df['cnpj'].astype(str).tolist()
        cnpjs_limpos = [limpar_cnpj(c) for c in cnpjs]

        # Prepara Excel (aplica estilos uma única vez)
        df.to_excel(caminho_excel, index=False)
        wb = load_workbook(caminho_excel)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
        wb.save(caminho_excel)
        wb.close()

        # Coleta pastas mês e configura progresso
        meses = coletar_pastas(pasta_pdf_base)
        total_steps = len(meses) * len(df)
        progress_callback('configure', maximum=total_steps)

        step = 0
        start_time = time.time()

        # Itera por mês com cache de PDFs
        for ano, mes_str, pasta_mes in meses:
            status_text.set(f"🚀 Processando {mes_str}")
            terminal.insert(tk.END, f"🚀 Processando {mes_str}\n")
            terminal.see(tk.END)
            pdfs = [os.path.join(pasta_mes, f) for f in os.listdir(pasta_mes) if f.lower().endswith('.pdf')]
            terminal.insert(tk.END, f"Mês {mes_str}: {len(pdfs)} PDFs encontrados\n")
            terminal.see(tk.END)

            textos_cache = carregar_textos_pdfs(pdfs)

            for idx, row in df.iterrows():
                while pause_event.is_set():
                    time.sleep(0.1)
                col = mes_str
                if str(row.get(col, '')).strip().lower() == 'concluído':
                    step += 1
                    progress_callback('step', value=1, start=start_time)
                    continue

                resultados = {}

                def busca_em_pdf(p):
                    encontrados = []
                    for i, texto in enumerate(textos_cache.get(p, [])):
                        if cnpjs[idx] in texto or cnpjs_limpos[idx] in limpar_cnpj(texto):
                            m, a = extrair_comp(texto)
                            if m and a:
                                encontrados.append((a, m, p, i))
                    return encontrados

                with ThreadPoolExecutor() as executor:
                    for found in executor.map(busca_em_pdf, pdfs):
                        for a, m, p, i in found:
                            resultados.setdefault((a, m), []).append((p, i))

                if resultados:
                    for (a, m), lst in resultados.items():
                        pasta_filial = os.path.join(pasta_destino, f"Filial - {row['Filial']}", ano)
                        os.makedirs(pasta_filial, exist_ok=True)
                        novo = fitz.open()
                        vistos = set()
                        for p, i in lst:
                            tex = textos_cache[p][i]
                            h = hash(tex)
                            if h in vistos:
                                continue
                            vistos.add(h)
                            novo.insert_pdf(fitz.open(p), from_page=i, to_page=i)
                        out = os.path.join(pasta_filial, f"SEFIP - {m}.pdf")
                        novo.save(out)
                        novo.close()
                        df.at[idx, mes_str] = 'Concluído'
                        terminal.insert(tk.END, f"{row['Filial']} - {ano}/{m} - OK\n")
                        terminal.see(tk.END)
                else:
                    df.at[idx, mes_str] = 'Não encontrado'

                step += 1
                progress_callback('step', value=1, start=start_time)

            # Salva Excel por mês
            df.to_excel(caminho_excel, index=False)

        status_text.set("🎉 Processo finalizado!")
        terminal.insert(tk.END, "🎉 Processo finalizado!\n")
        terminal.see(tk.END)
    except Exception as e:
        status_text.set(f"❌ Erro: {str(e)}")
        terminal.insert(tk.END, f"❌ Erro: {str(e)}\n")
        terminal.see(tk.END)


def iniciar_interface():
    style = Style(theme='vapor')
    root = style.master
    root.title("SEFIP")
    root.geometry("800x600")

    # Variables
    vars_vars = {
        'excel': tk.StringVar(),
        'base': tk.StringVar(),
        'dest': tk.StringVar()
    }
    status = tk.StringVar(value="Aguardando início...")
    pause_event = threading.Event()
    total = tk.IntVar()
    
    # Create widgets
    top = ttk.Frame(root, padding=10)
    top.pack(fill='x')
    
    def selecione(v, ftypes=None):
        if v == 'excel':
            path = filedialog.askopenfilename(filetypes=ftypes)
        else:
            path = filedialog.askdirectory()
        vars_vars[v].set(path)

    # Create input fields
    campos = [
        ('Excel', 'excel', [('Excel', '*.xlsx')]),
        ('Pasta Base', 'base', None),
        ('Pasta Destino', 'dest', None)
    ]
    
    for i, (rotulo, key, ftypes) in enumerate(campos):
        ttk.Label(top, text=f"{rotulo}:").grid(row=i, column=0, sticky='w')
        ttk.Entry(top, textvariable=vars_vars[key], width=60).grid(row=i, column=1)
        ttk.Button(top, text="Selecionar", command=lambda k=key, ft=ftypes: selecione(k, ft)).grid(row=i, column=2)

    # Progress bar
    bar = ttk.Progressbar(root, orient='horizontal', length=750, mode='determinate')
    bar.pack(pady=5)
    
    # Info frame
    info_frame = ttk.Frame(root)
    info_frame.pack(fill='x')
    pct = ttk.Label(info_frame, text="0%")
    pct.pack(side='left')
    eta_lbl = ttk.Label(info_frame, text="ETA: --:--:--")
    eta_lbl.pack(side='left', padx=10)
    
    # Terminal/log area
    log_frame = ttk.Frame(root)
    log_frame.pack(fill='both', expand=True)
    terminal = scrolledtext.ScrolledText(log_frame, height=15)
    terminal.pack(fill='both', expand=True, padx=5, pady=5)

    def progress_callback(action, **kwargs):
        if action == 'configure':
            bar.config(maximum=kwargs.get('maximum', 1))
            total.set(kwargs.get('maximum', 1))
        elif action == 'step':
            current = bar['value'] + kwargs.get('value', 1)
            bar['value'] = current
            pct.config(text=f"{current/total.get()*100:.1f}%")
            
            # Calculate ETA
            elapsed = time.time() - kwargs.get('start', time.time())
            if current > 0:
                remaining = (elapsed / current) * (total.get() - current)
                eta_str = time.strftime("%H:%M:%S", time.gmtime(remaining))
                eta_lbl.config(text=f"ETA: {eta_str}")

    # Control buttons
    btn_frame = ttk.Frame(root)
    btn_frame.pack(fill='x', pady=5)
    
    def toggle_pause():
        if pause_event.is_set():
            pause_event.clear()
            btn_pause.config(text="Pausar")
        else:
            pause_event.set()
            btn_pause.config(text="Continuar")

    btn_pause = ttk.Button(btn_frame, text="Pausar", command=toggle_pause)
    btn_pause.pack(side='left', padx=5)

    def start():
        if not all(v.get() for v in vars_vars.values()):
            messagebox.showerror("Erro", "Selecione todos os caminhos antes de iniciar.")
            return
        
        # Clear terminal
        terminal.delete('1.0', tk.END)
        
        # Start processing thread
        threading.Thread(
            target=processar,
            args=(
                vars_vars['excel'].get(),
                vars_vars['base'].get(),
                vars_vars['dest'].get(),
                status,
                progress_callback,
                pause_event,
                terminal
            ),
            daemon=True
        ).start()

    btn_start = ttk.Button(btn_frame, text="Iniciar", command=start)
    btn_start.pack(side='left', padx=5)
    
    # Status label
    ttk.Label(root, textvariable=status).pack(side='bottom', pady=5)

    root.mainloop()


if __name__ == '__main__':
    iniciar_interface()