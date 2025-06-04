import os
import re
import pandas as pd
import fitz  # PyMuPDF for handling PDF files
from concurrent.futures import ThreadPoolExecutor
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from ttkbootstrap import Style, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def limpar_cnpj(cnpj):
    """
    Remove todos os caracteres não numéricos de um CNPJ.
    """
    return re.sub(r'\D', '', str(cnpj))


def extrair_comp(texto):
    """
    Extrai mês e ano de competência a partir do texto, suportando ambos os formatos.
    """
    # Formato "COMP: MM/AAAA"
    match_comp = re.search(r'COMP\s*[:\.\-]?\s*(\d{2})/(\d{4})', texto)
    if match_comp:
        return (match_comp.group(1), match_comp.group(2))

    # Formato "Comp. Apuração: MM/AAAA"
    match_apuracao = re.search(r'Comp\.\s*Apuração\s*(\d{2})/(\d{4})', texto)
    if match_apuracao:
        return (match_apuracao.group(1), match_apuracao.group(2))

    return (None, None)


def coletar_pastas(pasta_base):
    """
    Retorna lista de tuplas (ano, mes_str, caminho_pasta) para todas pastas mês existentes.
    """
    pastas = []
    for ano in os.listdir(pasta_base):
        dir_ano = os.path.join(pasta_base, ano)
        if not os.path.isdir(dir_ano):
            continue
        for mes in range(1, 13):
            mes_str = f"{mes:02d}_{ano}"
            path = os.path.join(dir_ano, mes_str)
            if os.path.isdir(path):
                pastas.append((ano, mes_str, path))
    return pastas


def carregar_textos_pdfs(pdfs):
    """
    Lê todas as páginas de cada PDF uma única vez em cache.
    Retorna dict: {pdf_path: [texto_página0, texto_página1, ...]}
    """
    cache = {}
    for p in pdfs:
        try:
            doc = fitz.open(p)
            cache[p] = [doc.load_page(i).get_text() or "" for i in range(len(doc))]
            doc.close()
        except Exception:
            cache[p] = []
    return cache


def processar(caminho_excel, pasta_pdf_base, pasta_destino, status_text, progress_callback, pause_event, terminal):
    try:
        os.makedirs(pasta_destino, exist_ok=True)
        status_text.set("📚 Lendo lista de CNPJs...")
        terminal.insert(tk.END, "📚 Lendo lista de CNPJs...\n")
        terminal.see(tk.END)

        # Carrega Excel
        df = pd.read_excel(caminho_excel)
        cnpjs_para_buscar = df['cnpj'].astype(str).tolist()
        cnpjs_limpos_para_buscar = [limpar_cnpj(c) for c in cnpjs_para_buscar]

        # Prepara Excel (aplica estilos uma única vez)
        df.to_excel(caminho_excel, index=False)
        wb = load_workbook(caminho_excel)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
        wb.save(caminho_excel)
        wb.close()

        # Coleta pastas mês e configura progresso
        meses = coletar_pastas(pasta_pdf_base)
        total_steps = len(meses) * len(df)
        progress_callback('configure', maximum=total_steps)

        step = 0
        start_time = time.time()

        # Itera por mês com cache de PDFs
        for ano_pasta, mes_str_pasta, pasta_mes in meses:
            status_text.set(f"🚀 Processando {mes_str_pasta}")
            terminal.insert(tk.END, f"🚀 Processando {mes_str_pasta}\n")
            terminal.see(tk.END)
            pdfs = [os.path.join(pasta_mes, f) for f in os.listdir(pasta_mes) if f.lower().endswith('.pdf')]
            terminal.insert(tk.END, f"Mês {mes_str_pasta}: {len(pdfs)} PDFs encontrados\n")
            terminal.see(tk.END)

            textos_cache = carregar_textos_pdfs(pdfs)

            for idx, row in df.iterrows():
                while pause_event.is_set():
                    time.sleep(0.1)
                col_excel_competencia = mes_str_pasta
                
                # Se já está como 'Concluído' no Excel, pula.
                if str(row.get(col_excel_competencia, '')).strip().lower() == 'concluído':
                    step += 1
                    progress_callback('step', value=1, start=start_time)
                    continue

                cnpj_excel = cnpjs_para_buscar[idx]
                cnpj_limpo_excel = cnpjs_limpos_para_buscar[idx]

                # Lista para armazenar TODAS as páginas encontradas para este CNPJ e competência
                paginas_encontradas = []
                vistos_hash = set() # Para evitar duplicatas de páginas (mesmo conteúdo)

                # Itera pelos PDFs na pasta do mês
                for p_path in pdfs:
                    paginas_texto = textos_cache.get(p_path, [])
                    for i, texto_pagina in enumerate(paginas_texto):
                        mes_comp, ano_comp = extrair_comp(texto_pagina)
                        
                        # Verifica se a competência do PDF é a mesma da pasta que estamos processando
                        if mes_comp == mes_str_pasta[:2] and ano_comp == mes_str_pasta[3:]:
                            
                            # Tenta encontrar o CNPJ completo do Excel na página
                            # Usamos re.escape para tratar caracteres especiais como pontos e barras
                            if re.search(re.escape(cnpj_limpo_excel), limpar_cnpj(texto_pagina)):
                                # Adiciona à lista se a página não foi adicionada antes (usando hash do texto)
                                current_page_hash = hash(texto_pagina)
                                if current_page_hash not in vistos_hash:
                                    paginas_encontradas.append((p_path, i))
                                    vistos_hash.add(current_page_hash)
                                continue # Continua para a próxima página do mesmo PDF, pode haver outras ocorrências

                            # Se não encontrou o CNPJ completo, tenta encontrar a base do CNPJ (primeiros 8 dígitos)
                            cnpj_base_excel = cnpj_limpo_excel[:8]
                            if re.search(re.escape(cnpj_base_excel), limpar_cnpj(texto_pagina)):
                                # Adiciona à lista se a página não foi adicionada antes (usando hash do texto)
                                current_page_hash = hash(texto_pagina)
                                if current_page_hash not in vistos_hash:
                                    paginas_encontradas.append((p_path, i))
                                    vistos_hash.add(current_page_hash)
                                continue # Continua para a próxima página do mesmo PDF

                # Se encontrou UMA OU MAIS páginas para o CNPJ e competência atual
                if paginas_encontradas:
                    pasta_filial = os.path.join(pasta_destino, f"Filial - {row['Filial']}", ano_pasta)
                    os.makedirs(pasta_filial, exist_ok=True)
                    novo_pdf_combinado = fitz.open() # Objeto PDF para o novo arquivo combinado

                    for p_original_path, page_index in paginas_encontradas:
                        try:
                            original_doc = fitz.open(p_original_path)
                            novo_pdf_combinado.insert_pdf(original_doc, from_page=page_index, to_page=page_index)
                            original_doc.close()
                        except Exception as page_err:
                            terminal.insert(tk.END, f"⚠️ Erro ao inserir página {page_index} de {os.path.basename(p_original_path)}: {page_err}\n")
                            terminal.see(tk.END)
                            # Se uma página falhar, tenta as outras

                    # Salva o novo PDF combinado
                    out = os.path.join(pasta_filial, f"SEFIP - {mes_str_pasta[:2]}.pdf")
                    novo_pdf_combinado.save(out)
                    novo_pdf_combinado.close()
                    
                    df.at[idx, col_excel_competencia] = 'Concluído'
                    terminal.insert(tk.END, f"{row['Filial']} - {ano_pasta}/{mes_str_pasta[:2]} - OK. Páginas extraídas: {len(paginas_encontradas)}\n")
                    terminal.see(tk.END)
                else:
                    df.at[idx, col_excel_competencia] = 'Não encontrado'
                    terminal.insert(tk.END, f"{row['Filial']} - {ano_pasta}/{mes_str_pasta[:2]} - Não encontrado.\n")
                    terminal.see(tk.END)

                step += 1
                progress_callback('step', value=1, start=start_time)

            # Salva Excel por mês
            df.to_excel(caminho_excel, index=False)

        status_text.set("🎉 Processo finalizado!")
        terminal.insert(tk.END, "🎉 Processo finalizado!\n")
        terminal.see(tk.END)
    except Exception as e:
        status_text.set(f"❌ Erro: {str(e)}")
        terminal.insert(tk.END, f"❌ Erro: {str(e)}\n")
        terminal.see(tk.END)


def iniciar_interface():
    style = Style(theme='vapor')
    root = style.master
    root.title("SEFIP")
    root.geometry("800x600")

    # Variables
    vars_vars = {
        'excel': tk.StringVar(),
        'base': tk.StringVar(),
        'dest': tk.StringVar()
    }
    status = tk.StringVar(value="Aguardando início...")
    pause_event = threading.Event()
    total = tk.IntVar()
    
    # Create widgets
    top = ttk.Frame(root, padding=10)
    top.pack(fill='x')
    
    def selecione(v, ftypes=None):
        if v == 'excel':
            path = filedialog.askopenfilename(filetypes=ftypes)
        else:
            path = filedialog.askdirectory()
        vars_vars[v].set(path)

    # Create input fields
    campos = [
        ('Excel', 'excel', [('Excel', '*.xlsx')]),
        ('Pasta Base', 'base', None),
        ('Pasta Destino', 'dest', None)
    ]
    
    for i, (rotulo, key, ftypes) in enumerate(campos):
        ttk.Label(top, text=f"{rotulo}:").grid(row=i, column=0, sticky='w')
        ttk.Entry(top, textvariable=vars_vars[key], width=60).grid(row=i, column=1)
        ttk.Button(top, text="Selecionar", command=lambda k=key, ft=ftypes: selecione(k, ft)).grid(row=i, column=2)

    # Progress bar
    bar = ttk.Progressbar(root, orient='horizontal', length=750, mode='determinate')
    bar.pack(pady=5)
    
    # Info frame
    info_frame = ttk.Frame(root)
    info_frame.pack(fill='x')
    pct = ttk.Label(info_frame, text="0%")
    pct.pack(side='left')
    eta_lbl = ttk.Label(info_frame, text="ETA: --:--:--")
    eta_lbl.pack(side='left', padx=10)
    
    # Terminal/log area
    log_frame = ttk.Frame(root)
    log_frame.pack(fill='both', expand=True)
    terminal = scrolledtext.ScrolledText(log_frame, height=15)
    terminal.pack(fill='both', expand=True, padx=5, pady=5)

    def progress_callback(action, **kwargs):
        if action == 'configure':
            bar.config(maximum=kwargs.get('maximum', 1))
            total.set(kwargs.get('maximum', 1))
        elif action == 'step':
            current = bar['value'] + kwargs.get('value', 1)
            bar['value'] = current
            pct.config(text=f"{current/total.get()*100:.1f}%")
            
            # Calculate ETA
            elapsed = time.time() - kwargs.get('start', time.time())
            if current > 0:
                remaining = (elapsed / current) * (total.get() - current)
                eta_str = time.strftime("%H:%M:%S", time.gmtime(remaining))
                eta_lbl.config(text=f"ETA: {eta_str}")

    # Control buttons
    btn_frame = ttk.Frame(root)
    btn_frame.pack(fill='x', pady=5)
    
    def toggle_pause():
        if pause_event.is_set():
            pause_event.clear()
            btn_pause.config(text="Pausar")
        else:
            pause_event.set()
            btn_pause.config(text="Continuar")

    btn_pause = ttk.Button(btn_frame, text="Pausar", command=toggle_pause)
    btn_pause.pack(side='left', padx=5)

    def start():
        if not all(v.get() for v in vars_vars.values()):
            messagebox.showerror("Erro", "Selecione todos os caminhos antes de iniciar.")
            return
        
        # Clear terminal
        terminal.delete('1.0', tk.END)
        
        # Start processing thread
        threading.Thread(
            target=processar,
            args=(
                vars_vars['excel'].get(),
                vars_vars['base'].get(),
                vars_vars['dest'].get(),
                status,
                progress_callback,
                pause_event,
                terminal
            ),
            daemon=True
        ).start()

    btn_start = ttk.Button(btn_frame, text="Iniciar", command=start)
    btn_start.pack(side='left', padx=5)
    
    # Status label
    ttk.Label(root, textvariable=status).pack(side='bottom', pady=5)

    root.mainloop()


if __name__ == '__main__':
    iniciar_interface()