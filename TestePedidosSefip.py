import ttkbootstrap as tb
from ttkbootstrap.constants import *
import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Cria a janela principal
app = tb.Window(themename="darkly")
app.title("Pedidos SEFIP")
app.geometry("900x680")

# Lista para armazenar os dados das filiais
filiais_data = []

# ======= Funções de formatação =======

def formatar_cnpj(event):
    """
    Formata o CNPJ conforme o padrão "XX.XXX.XXX/XXXX-XX".
    """
    texto = re.sub(r'\D', '', cnpj_entry.get())
    formatado = ""
    if len(texto) > 0:
        formatado += texto[:2]
    if len(texto) > 2:
        formatado += "." + texto[2:5]
    if len(texto) > 5:
        formatado += "." + texto[5:8]
    if len(texto) > 8:
        formatado += "/" + texto[8:12]
    if len(texto) > 12:
        formatado += "-" + texto[12:14]
    cnpj_entry.delete(0, tk.END)
    cnpj_entry.insert(0, formatado)

def formatar_data(event):
    """
    Formata a data conforme o padrão "DD/MM/YYYY".
    """
    texto = re.sub(r'\D', '', data_entry.get())
    formatado = ""
    if len(texto) > 0:
        formatado += texto[:2]
    if len(texto) > 2:
        formatado += "/" + texto[2:4]
    if len(texto) > 4:
        formatado += "/" + texto[4:8]
    data_entry.delete(0, tk.END)
    data_entry.insert(0, formatado)

# ======= Entradas de dados =======

# Entrada de dados para filial, CNPJ, data e observações
filial_label = tb.Label(app, text="Filial:")
filial_label.pack(pady=5)
filial_entry = tb.Entry(app, width=40)
filial_entry.pack(pady=5)

cnpj_label = tb.Label(app, text="CNPJ:")
cnpj_label.pack(pady=5)
cnpj_entry = tb.Entry(app, width=40)
cnpj_entry.pack(pady=5)
cnpj_entry.bind("<KeyRelease>", formatar_cnpj)

data_label = tb.Label(app, text="Data de Entrega:")
data_label.pack(pady=5)
data_entry = tb.Entry(app, width=40)
data_entry.pack(pady=5)
data_entry.bind("<KeyRelease>", formatar_data)

obs_label = tb.Label(app, text="Observações:")
obs_label.pack(pady=5)
obs_entry = tb.Entry(app, width=80)
obs_entry.pack(pady=5)

# ======= Checkboxes dos anos =======

# Frame para os checkboxes dos anos
checkbox_frame = tb.Frame(app)
checkbox_frame.pack(pady=10)

# Criação dos anos (2009 a 2025) para o usuário marcar
anos = list(range(2009, 2026))
checkbox_vars = {}
for i, ano in enumerate(anos):
    var = tk.BooleanVar()
    checkbox = tb.Checkbutton(checkbox_frame, text=str(ano), variable=var, bootstyle="info")
    checkbox.grid(row=i // 9, column=i % 9, padx=5, pady=2, sticky="w")
    checkbox_vars[ano] = var

# ======= Lista de filiais adicionadas =======

# Exibe as filiais que foram adicionadas
lista_label = tb.Label(app, text="Filiais adicionadas:")
lista_label.pack(pady=2)
lista_box = tk.Listbox(app, height=6, width=100)
lista_box.pack(pady=5)

# ======= Funções principais =======

def adicionar_filial():
    """
    Adiciona uma nova filial à lista de dados e atualiza a lista na interface.
    """
    filial = filial_entry.get().strip()
    cnpj = cnpj_entry.get().strip()
    data_entrega = data_entry.get().strip()
    observacoes = obs_entry.get().strip()
    anos_marcados = [ano for ano, var in checkbox_vars.items() if var.get()]

    # Verifica se todos os campos obrigatórios foram preenchidos
    if not filial or not cnpj or not data_entrega:
        messagebox.showerror("Erro", "Preencha todos os campos obrigatórios: Filial, CNPJ e Data de Entrega.")
        return
    if not anos_marcados:
        messagebox.showerror("Erro", "Selecione pelo menos um ano.")
        return

    # Adiciona os dados da filial à lista de filiais
    dados = {
        "Filial": filial,
        "CNPJ": cnpj,
        "Data de Entrega": data_entrega,
        "Observações": observacoes
    }

    for ano in anos:
        dados[str(ano)] = "x" if ano in anos_marcados else ""

    filiais_data.append(dados)
    lista_box.insert(
        tk.END,
        f"{filial} | CNPJ: {cnpj} | Entrega: {data_entrega} | Anos: {', '.join(map(str, anos_marcados))} | Obs: {observacoes}"
    )

    # Limpar campos de entrada
    filial_entry.delete(0, tk.END)
    cnpj_entry.delete(0, tk.END)
    data_entry.delete(0, tk.END)
    obs_entry.delete(0, tk.END)
    for var in checkbox_vars.values():
        var.set(False)

def remover_filial():
    """
    Remove uma filial selecionada da lista de dados.
    """
    selecionado = lista_box.curselection()
    if not selecionado:
        messagebox.showerror("Erro", "Selecione uma filial na lista para remover.")
        return
    index = selecionado[0]
    lista_box.delete(index)
    del filiais_data[index]

def aplicar_formatacao_excel(caminho):
    """
    Aplica formatação personalizada na planilha gerada (cabeçalho, cores e bordas).
    """
    wb = load_workbook(caminho)
    ws = wb.active

    # Cores e estilos
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
    red_fill = PatternFill(start_color="CC3333", end_color="CC3333", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Formatação do cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border

    # Formatação das células
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.value == "x":
                cell.fill = green_fill
            elif cell.column_letter in [chr(65 + i) for i in range(4, len(anos)+4)]:  # colunas de ano
                cell.fill = red_fill

    wb.save(caminho)

def gerar_planilha():
    """
    Gera a planilha com os dados das filiais e aplica a formatação.
    """
    if not filiais_data:
        messagebox.showerror("Erro", "Nenhuma filial adicionada.")
        return

    # Seleciona o local para salvar a planilha
    caminho = filedialog.asksaveasfilename(
        title="Salvar planilha como...",
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )

    if not caminho:
        return

    try:
        # Cria um DataFrame com os dados das filiais
        df = pd.DataFrame(filiais_data)
        df.to_excel(caminho, index=False)

        # Aplica a formatação na planilha
        aplicar_formatacao_excel(caminho)

        messagebox.showinfo("Sucesso", f"Planilha gerada com sucesso:\n{caminho}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar a planilha:\n{e}")

# ======= Botões =======

# Frame para os botões de adicionar e remover filiais
botao_frame = tb.Frame(app)
botao_frame.pack(pady=10)

# Botões de adicionar e remover filiais
add_btn = tb.Button(botao_frame, text="Adicionar Filial", command=adicionar_filial, bootstyle="primary")
add_btn.grid(row=0, column=0, padx=5)

remove_btn = tb.Button(botao_frame, text="Remover Filial Selecionada", command=remover_filial, bootstyle="danger")
remove_btn.grid(row=0, column=1, padx=5)

# Botão para gerar a planilha
gerar_btn = tb.Button(app, text="Gerar Planilha", command=gerar_planilha, bootstyle="success")
gerar_btn.pack(pady=10)

# ======= Iniciar app =======
app.mainloop()
